"""Tests for src.reporter.excel_report."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from tests.conftest import MockRow

from src.reporter.excel_report import (
    _format_price,
    _format_ratio,
    _get_customer_status_ru,
    _get_parse_status_ru,
    generate_excel_report,
)


# ── Fixture helpers ──────────────────────────────────────────────────────────


def _interesting_row(**overrides) -> MockRow:
    defaults = {
        "tender_title": "Поставка серверов",
        "tender_url": "https://rostender.info/tender/100",
        "tender_price": 5_000_000.0,
        "customer_name": 'ООО "Ромашка"',
        "customer_inn": "1234567890",
        "total_historical": 10,
        "total_analyzed": 8,
        "low_competition_count": 6,
        "competition_ratio": 0.75,
        "source": "primary",
    }
    defaults.update(overrides)
    return MockRow(defaults)


def _customer_row(**overrides) -> MockRow:
    defaults = {
        "inn": "1234567890",
        "name": 'ООО "Ромашка"',
        "status": "analyzed",
        "total_tenders": 15,
        "active_tenders": 3,
        "completed_tenders": 12,
    }
    defaults.update(overrides)
    return MockRow(defaults)


def _protocol_row(**overrides) -> MockRow:
    defaults = {
        "tender_id": "t1",
        "customer_inn": "1234567890",
        "tender_status": "completed",
        "participants_count": 2,
        "parse_source": "html",
        "parse_status": "success",
        "doc_path": "/tmp/doc.html",
        "notes": None,
    }
    defaults.update(overrides)
    return MockRow(defaults)


def _result_row(**overrides) -> MockRow:
    defaults = {
        "customer_inn": "1234567890",
        "total_historical": 10,
        "total_analyzed": 8,
    }
    defaults.update(overrides)
    return MockRow(defaults)


# ── Unit tests for helper functions ──────────────────────────────────────────


class TestFormatPrice:
    def test_none(self) -> None:
        assert _format_price(None) == ""

    def test_zero(self) -> None:
        assert _format_price(0.0) == "0"

    def test_positive(self) -> None:
        assert _format_price(5_000_000.0) == "5,000,000"

    def test_small(self) -> None:
        assert _format_price(99.5) == "100"


class TestFormatRatio:
    def test_none(self) -> None:
        assert _format_ratio(None) == "N/A"

    def test_zero(self) -> None:
        assert _format_ratio(0.0) == "0%"

    def test_full(self) -> None:
        assert _format_ratio(1.0) == "100%"

    def test_fraction(self) -> None:
        assert _format_ratio(0.75) == "75%"


class TestGetCustomerStatusRu:
    def test_known_statuses(self) -> None:
        assert _get_customer_status_ru("new") == "Новый"
        assert _get_customer_status_ru("processing") == "В обработке"
        assert _get_customer_status_ru("analyzed") == "Проанализирован"
        assert _get_customer_status_ru("error") == "Ошибка"

    def test_unknown_status_passthrough(self) -> None:
        assert _get_customer_status_ru("unknown_xyz") == "unknown_xyz"


class TestGetParseStatusRu:
    def test_known_statuses(self) -> None:
        assert _get_parse_status_ru("success") == "Успешно"
        assert _get_parse_status_ru("failed") == "Ошибка"
        assert _get_parse_status_ru("skipped_scan") == "Пропущен (скан)"
        assert _get_parse_status_ru("no_protocol") == "Нет протокола"

    def test_unknown_status_passthrough(self) -> None:
        assert _get_parse_status_ru("something_else") == "something_else"


# ── Integration tests for generate_excel_report ──────────────────────────────


class TestGenerateExcelReport:
    """Integration tests that verify actual .xlsx file content."""

    def test_creates_file(self, tmp_path: Path) -> None:
        """Should create an xlsx file and return its path."""
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[_interesting_row()],
                all_results=[_result_row()],
                all_customers=[_customer_row()],
                all_protocols=[_protocol_row()],
            )

        assert filepath.exists()
        assert filepath.suffix == ".xlsx"
        assert filepath.parent == tmp_path

    def test_has_three_sheets(self, tmp_path: Path) -> None:
        """Report should have exactly 3 sheets."""
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[_interesting_row()],
                all_results=[_result_row()],
                all_customers=[_customer_row()],
                all_protocols=[_protocol_row()],
            )

        wb = load_workbook(filepath)
        assert len(wb.sheetnames) == 3
        assert "Интересные тендеры" in wb.sheetnames
        assert "Все заказчики" in wb.sheetnames
        assert "Детали анализа" in wb.sheetnames

    def test_interesting_sheet_content(self, tmp_path: Path) -> None:
        """Interesting sheet should contain header + data rows."""
        rows = [
            _interesting_row(tender_title="Тендер А"),
            _interesting_row(tender_title="Тендер Б", source="extended"),
        ]
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=rows,
                all_results=[_result_row()],
                all_customers=[_customer_row()],
                all_protocols=[_protocol_row()],
            )

        wb = load_workbook(filepath)
        ws = wb["Интересные тендеры"]
        # Header row + 2 data rows
        data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        assert len(data_rows) == 3  # 1 header + 2 data
        # Check header
        assert data_rows[0][0] == "Название"
        assert data_rows[0][4] == "ИНН"
        # Check first data row
        assert data_rows[1][0] == "Тендер А"
        assert data_rows[1][9] == "Основной"
        # Check second data row
        assert data_rows[2][0] == "Тендер Б"
        assert data_rows[2][9] == "Расширенный"

    def test_customers_sheet_content(self, tmp_path: Path) -> None:
        """Customers sheet should contain customer data."""
        customers = [
            _customer_row(inn="111", name="First"),
            _customer_row(inn="222", name="Second", status="new"),
        ]
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[],
                all_customers=customers,
                all_protocols=[],
            )

        wb = load_workbook(filepath)
        ws = wb["Все заказчики"]
        data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        assert len(data_rows) == 3  # 1 header + 2 data
        assert data_rows[1][0] == "111"
        assert data_rows[1][1] == "First"
        assert data_rows[1][2] == "Проанализирован"
        assert data_rows[2][0] == "222"
        assert data_rows[2][2] == "Новый"

    def test_analysis_details_sheet(self, tmp_path: Path) -> None:
        """Analysis details sheet should contain protocol data."""
        protocols = [
            _protocol_row(
                tender_id="t1", participants_count=3, tender_status="completed"
            ),
            _protocol_row(
                tender_id="t2", participants_count=None, parse_status="failed"
            ),
        ]
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[_result_row()],
                all_customers=[],
                all_protocols=protocols,
            )

        wb = load_workbook(filepath)
        ws = wb["Детали анализа"]
        data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        assert len(data_rows) == 3  # 1 header + 2 data
        assert data_rows[1][0] == "t1"
        assert data_rows[1][2] == "Завершён"
        assert data_rows[1][3] == 3
        assert data_rows[2][0] == "t2"
        assert data_rows[2][3] == "N/A"
        assert data_rows[2][5] == "Ошибка"

    def test_empty_data(self, tmp_path: Path) -> None:
        """Should produce valid xlsx even with all empty inputs."""
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[],
                all_customers=[],
                all_protocols=[],
            )

        assert filepath.exists()
        wb = load_workbook(filepath)
        assert len(wb.sheetnames) == 3

    def test_protocol_without_tender_status(self, tmp_path: Path) -> None:
        """Protocol row missing tender_status should fallback to 'Активный'."""
        protocol = MockRow(
            {
                "tender_id": "t99",
                "customer_inn": "999",
                "participants_count": 1,
                "parse_source": "html",
                "parse_status": "success",
                "doc_path": None,
                "notes": None,
            }
        )
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[],
                all_customers=[],
                all_protocols=[protocol],
            )

        wb = load_workbook(filepath)
        ws = wb["Детали анализа"]
        data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        assert data_rows[1][2] == "Активный"

    def test_filename_contains_timestamp(self, tmp_path: Path) -> None:
        """Generated filename should contain a timestamp pattern."""
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[],
                all_customers=[],
                all_protocols=[],
            )

        assert filepath.name.startswith("report_")
        # Pattern: report_YYYY-MM-DD_HH-MM.xlsx
        parts = filepath.stem.split("_")
        assert len(parts) >= 3

    def test_customer_none_name(self, tmp_path: Path) -> None:
        """Customer with name=None should show empty string."""
        customers = [_customer_row(name=None)]
        with patch("src.reporter.excel_report.REPORTS_DIR", tmp_path):
            filepath = generate_excel_report(
                interesting_results=[],
                all_results=[],
                all_customers=customers,
                all_protocols=[],
            )

        wb = load_workbook(filepath)
        ws = wb["Все заказчики"]
        data_rows = list(ws.iter_rows(min_row=1, values_only=True))
        # name column should be empty string (which openpyxl reads as None for empty cells)
        assert data_rows[1][1] in ("", None)
