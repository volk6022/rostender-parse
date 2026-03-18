"""Генерация Excel-отчёта по активным тендерам."""

from __future__ import annotations

from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment

from src.config import REPORTS_DIR
from src.reporter.excel_report import HEADER_FILL, HEADER_FONT, _format_price


def generate_active_tenders_report(
    tenders: Sequence[Any],
    session_id: str | None = None,
    stage: str | None = None,
) -> Path | None:
    """Сгенерировать Excel со списком активных тендеров."""
    if not tenders:
        logger.warning("Нет активных тендеров для генерации отчёта")
        return None

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Активные тендеры")
    else:
        ws.title = "Активные тендеры"

    headers = [
        "ID тендера",
        "Название",
        "Цена (₽)",
        "ИНН заказчика",
        "Ссылка",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for t in tenders:
        ws.append(
            [
                t["tender_id"],
                t["title"] or "",
                _format_price(t["price"]),
                t["customer_inn"],
                t["url"] or "",
            ]
        )

    # Автоподбор ширины колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 60)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Формируем уникальное имя файла
    if session_id:
        base = f"report_{session_id}"
        if stage:
            base += f"_{stage}"
        base += ".xlsx"
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        base = f"active_tenders_{timestamp}.xlsx"

    filepath = REPORTS_DIR / base

    # Атомарная запись
    tmp_filepath = REPORTS_DIR / f"{base}.tmp"
    try:
        wb.save(tmp_filepath)
        tmp_filepath.replace(filepath)
    except Exception:
        if tmp_filepath.exists():
            tmp_filepath.unlink()
        raise

    logger.success("Отчёт по активным тендерам сохранён: {}", filepath)
    return filepath
