"""Генерация Excel-отчёта (.xlsx)."""

from __future__ import annotations

from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import REPORTS_DIR


HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INTERESTING_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)


def _format_price(price: float | None) -> str:
    if price is None:
        return ""
    return f"{price:,.0f}"


def _format_ratio(ratio: float | None) -> str:
    if ratio is None:
        return "N/A"
    return f"{ratio:.0%}"


def _get_customer_status_ru(status: str) -> str:
    status_map = {
        "new": "Новый",
        "processing": "В обработке",
        "extended_processing": "Расширенный поиск",
        "extended_analyzed": "Расширенный анализ завершён",
        "analyzed": "Проанализирован",
        "error": "Ошибка",
    }
    return status_map.get(status, status)


def _get_parse_status_ru(status: str) -> str:
    status_map = {
        "success": "Успешно",
        "failed": "Ошибка",
        "skipped_scan": "Пропущен (скан)",
        "no_protocol": "Нет протокола",
    }
    return status_map.get(status, status)


def generate_excel_report(
    interesting_results: Sequence[Any],
    all_results: Sequence[Any],
    all_customers: Sequence[Any],
    all_protocols: Sequence[Any],
) -> Path:
    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _write_interesting_sheet(wb, interesting_results)
    _write_customers_sheet(wb, all_customers)
    _write_analysis_details_sheet(wb, all_protocols, all_results)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"report_{timestamp}.xlsx"
    filepath = REPORTS_DIR / filename

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)

    logger.success("Excel-отчёт сохранён: {}", filepath)
    return filepath


def _write_interesting_sheet(wb: Workbook, results: Sequence[Any]) -> None:
    ws = wb.create_sheet("Интересные тендеры")

    headers = [
        "Название",
        "URL",
        "Цена (₽)",
        "Заказчик",
        "ИНН",
        "Всего исторических",
        "Проанализировано",
        "Низкая конкуренция",
        "Доля",
        "Источник",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in results:
        ws.append(
            [
                row["tender_title"] or "",
                row["tender_url"] or "",
                _format_price(row["tender_price"]),
                row["customer_name"] or "",
                row["customer_inn"],
                row["total_historical"],
                row["total_analyzed"],
                row["low_competition_count"],
                _format_ratio(row["competition_ratio"]),
                "Основной" if row["source"] == "primary" else "Расширенный",
            ]
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 60)


def _write_customers_sheet(wb: Workbook, customers: Sequence[Any]) -> None:
    ws = wb.create_sheet("Все заказчики")

    headers = ["ИНН", "Название", "Статус", "Всего тендеров", "Активных", "Завершённых"]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in customers:
        ws.append(
            [
                row["inn"],
                row["name"] or "",
                _get_customer_status_ru(row["status"]),
                row["total_tenders"] or 0,
                row["active_tenders"] or 0,
                row["completed_tenders"] or 0,
            ]
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)


def _write_analysis_details_sheet(
    wb: Workbook,
    protocols: Sequence[Any],
    results: Sequence[Any],
) -> None:
    ws = wb.create_sheet("Детали анализа")

    results_by_customer = {}
    for r in results:
        inn = r["customer_inn"]
        if inn not in results_by_customer:
            results_by_customer[inn] = []
        results_by_customer[inn].append(r)

    headers = [
        "tender_id",
        "Заказчик (ИНН)",
        "Статус тендера",
        "Участников",
        "Источник парсинга",
        "Статус парсинга",
        "Путь к документу",
        "Заметки",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in protocols:
        customer_inn = row["customer_inn"]
        try:
            tender_status = row["tender_status"]
        except (IndexError, KeyError):
            tender_status = ""

        ws.append(
            [
                row["tender_id"],
                customer_inn,
                "Завершён" if tender_status == "completed" else "Активный",
                row["participants_count"]
                if row["participants_count"] is not None
                else "N/A",
                row["parse_source"] or "",
                _get_parse_status_ru(row["parse_status"]),
                row["doc_path"] or "",
                row["notes"] or "",
            ]
        )

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
